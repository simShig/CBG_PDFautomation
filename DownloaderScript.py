import atexit
import time
import openpyxl
import os
import logging
import datetime
import requests
import pandas as pd
from bs4 import BeautifulSoup

#
#
excel_file_path = fr'C:\Users\Simon\Desktop\CBG\ReadingTaskAll.xlsx'
output_dir_path = fr'C:\Users\Simon\Desktop\CBG\PDFs'
logs_dir_path = fr'C:\Users\Simon\Desktop\CBG\LogFiles'
oxyUser = "RaCheck"
oxyPW = "RaCheck123"

proxies = None
log_file = None
blocked_sites = ["sciencedirect","ieeexplore", "aiia.csd", "link.springer","onlinelibrary.wiley"]  # Add more sites as needed -  "researchgate"

isNameDotPDF = True  ##distinguish between urls .../pdf/.. (FALSE) or ....pdf (true)

# indices = [0,1,2,4,7]



def startProxy(username, password):
    from requests.packages.urllib3.exceptions import InsecureRequestWarning
    requests.packages.urllib3.disable_warnings(InsecureRequestWarning)
    global proxies
    prxStr = f'http://{username}:{password}@pr.oxylabs.io:7777'
    proxy = prxStr
    proxies = {
        'http': proxy,
        'https': proxy
    }
    prntL(f'\tstarting proxy: {proxy}')


def getBibtex(row, rowNum, name_index,last_name_index, attackDefence_index, title_index, bibtex_index, pdfLink_index):  # rowNum > bibtex
    prntL(f"row num: {rowNum} [by {row[name_index]} {row[last_name_index]}, article name: {row[attackDefence_index]} ~ {row[title_index]}] ")
    bibtex = row[bibtex_index]
    if bibtex is None:
        bibtex = row[title_index]
    if bibtex is None:
        bibtex = row[pdfLink_index]
    return bibtex


def checkResponseStatus(response):
    logging.info(f'response code is: {response.status_code}')
    if response.status_code == 403:
        prntL("Bot blocked (403 Forbidden)")
    if response.status_code == 418:
        prntL("418 teapot <=> Bot blocked (like 403 Forbidden)")
    if response.status_code == 550:
        prntL("The recipient is blocking your email on the recipient's email server (550 Forbidden)")
    if "captcha" in response.text.lower():
        prntL("\t!!! You are facing RECAPTCHA !!!")
        return True
    return False


def find_pdf_links(bibtex):  # bibtex > pdfUrl
    # Make a request to Google Scholar with the 'bibtex' variable
    page = f'https://scholar.google.com/scholar?q={bibtex}'
    if proxies:
        response = requests.request('GET', page, verify=False, proxies=proxies)  # , proxies=proxies, verify=False)
        if checkResponseStatus(response):   ##if i need to try again because i got CAPTHCA
            prntL("second try after RECAPTCHA (bad proxy):")
            response = requests.request('GET', page, verify=False, proxies=proxies)  # , proxies=proxies, verify=False)
    else:
        response = requests.get(page)
        checkResponseStatus(response)
    if response.status_code == 200:
        soup = BeautifulSoup(response.text, 'html.parser')
        selectors = ['a[href$=".pdf"]', 'a[href*="/pdf"]', 'a[href*="/article/"]', 'a[href*="/download"]']
        for selector in selectors:
            pdf_links = soup.select(selector)
            if pdf_links:
                isNameDotPDF = selector in ('a[href$=".pdf"]',)
                return pdf_links, isNameDotPDF

        prntL("Could not find PDF URLs")
        # prntL(response.text)
        # logging.info("Could not find PDF URLs")
        return [], False


def markAsDownloaded(wasDownloaded, pdfUrl, rNum, sheet, pdfLinkIndex, wasDownloadedIndex):
    needVpn = False

    # Assuming that the tuple index 21 corresponds to column "V"
    sheet.cell(row=rNum, column=pdfLinkIndex+1, value=pdfUrl)  # article PDF link

    if wasDownloaded:
        sheet.cell(row=rNum, column=wasDownloadedIndex+1, value="V")  # wasDownloaded?

    # if any(site in pdfUrl.lower() for site in blocked_sites):
    #     needVpn = True
    #     sheet.cell(row=rNum, column=24, value="V")  # is handled by sci-hub?
    # prntL(f'\t\tMarkedInEXCEL: wasDownloaded?{wasDownloaded}, need VPN? {needVpn} ')
    global wb
    wb.save(excel_file_path)


def twistBlockedUrl(pdfUrl):        ##add special twists if encounter mor problematiqe domains.
    newUrl = 'https://sci-hub.se/'
    if "ieeexplore" in pdfUrl.lower():
        base_name = os.path.splitext(os.path.basename(pdfUrl))[0]
        # Extract the desired part from the base name
        desired_part_with_zeros = base_name.split('/')[-1]
        # Remove leading zeros
        desired_part_without_zeros = str(int(desired_part_with_zeros))
        newUrl = 'https://sci-hub.se/https://ieeexplore.ieee.org/document/'+desired_part_without_zeros  #apply changes for IEEExplore
    else:
        newUrl = newUrl+pdfUrl
    print(f'\tnew URL is: {newUrl}')
    return newUrl

def download_pdf(pdfUrl, articleName, rNum):  # returns 1 if downloaded, -1 if failed
    if articleName==None:   #in case no title was given
        articleName = "NoArticleTitle"
    articleName = ''.join(char for char in articleName if char.isalpha())  # remove non-alphabetical chars
    articleName = f'{rNum}_{articleName}'
    prntL("\t\t" + articleName)
    os.makedirs(output_dir_path, exist_ok=True)
    # Define the path to the folder where you want to save the PDF
    folder_path = output_dir_path + '\\'

    # Combine the folder path and the filename to create the full file path
    file_path = folder_path + articleName + '.pdf'
    tryWithSciHub = False       #when managing knowen issues
    # Send a GET request to the URL
    # if proxies:
    #     urlResp = requests.request('GET', pdfUrl, proxies=proxies, verify=False)  # urlResponse
    # else:
    try:
        if any(site in pdfUrl.lower() for site in blocked_sites):
            tryWithSciHub = True
            pdfUrl = twistBlockedUrl(pdfUrl)
            prntL("@@@@running a try with SciHub")
        # Send a GET request to the URL with a timeout
        urlResp = requests.get(pdfUrl, timeout=4)  # timout (secs) for loading the pdfUrl

        # Check if the request was successful (status code 200)
        if urlResp.status_code == 200:
            if tryWithSciHub:
                soup = BeautifulSoup(urlResp.content, 'html.parser')

                if "Unfortunately, Sci-Hub" in urlResp.text:
                    prntL(f'\n\t[X] !!!~An error occurred while downloading PDF on row {rNum}:\n'
                          f'the link is not available on sci-hub')
                    return -1

                # Find the <embed> tag within the <div id="article"> section
                embed_tag = soup.find('div', id='article').find('embed')

                # Extract the PDF URL from the src attribute of the <embed> tag
                pdf_url_relative = embed_tag['src']

                # Create the absolute PDF URL by concatenating with the base URL
                base_url = 'https://sci-hub.se/'
                pdf_url = base_url + pdf_url_relative

                # Download the PDF using the extracted URL
                pdf_resp = requests.get(pdf_url, timeout=10)
                if pdf_resp.status_code == 200:
                    urlResp=pdf_resp
                else:
                    prntL(f'\n\t[X] !!!~An error occurred while downloading PDF on row {rNum}:\n'
                          f'the pdfResp status code (from SciHub) is: {pdf_resp.status_code}')
                    return -1

            # Open the file in binary write mode and write the content
            with open(file_path, 'wb') as f:
                f.write(urlResp.content)

            # Check the file size
            file_size = os.path.getsize(file_path)
            if file_size < 100:
                prntL(
                    f'\n\t[X] ~ Downloaded file on row {rNum} is too small (size: {file_size} bytes). Possible corruption.\n')
                return -1
            else:
                prntL(f'\t[V] PDF saved to: {file_path}')
                return 1
        else:
            prntL(f'\n\t[X] !!!~Failed to download PDF on row {rNum}. Status code: {urlResp.status_code}. For further examination check the log.\n')
            logging.info(urlResp.text)
            return -1
    except requests.Timeout:
        prntL(f'\n\t[X] !!!~Request timed out while downloading PDF on row {rNum}.\n')
        return -1
    except requests.RequestException as e:
        prntL(f'\n\t[X] !!!~An error occurred while downloading PDF on row {rNum}: {e}\n')
        return -1


def setup_logging():
    global log_file
    os.makedirs(logs_dir_path, exist_ok=True)

    # Generate a unique log file name with a timestamp
    log_str = f"log_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.log"
    log_file = os.path.join(logs_dir_path + '\\', log_str)
    print(log_file)
    logging.basicConfig(filename=log_file, level=logging.INFO, format='%(asctime)s [%(levelname)s] - %(message)s')


def close_logging():
    global log_file
    if log_file is not None:
        log_file_handler = None
        for handler in logging.getLogger().handlers:
            if isinstance(handler, logging.FileHandler):
                log_file_handler = handler
                break

        if log_file_handler is not None:
            log_file_handler.stream.close()
            logging.getLogger().removeHandler(log_file_handler)
        log_file = None


def prntL(someString):
    logging.info(someString)
    print(someString)


def cleanup():
    prntL("~~atExit: in cleanup")
    global wb
    # if 'wb' in locals():
    wb.save(excel_file_path)
    prntL("~~atExit: saving WorkBook")
    wb.close()
    prntL("~~atExit: WorkBook closed")
    close_logging()


# Register the cleanup function with atexit
atexit.register(cleanup)


def needToDownloadRow(isDownloadedCell,PDFlinkCell):
    if isDownloadedCell == "V":  #isDownloaded
        return False
    if PDFlinkCell == None:   #no link yet
        return True
    return True

def initializeIndices(sheet):
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    column_indices = {name: index for index, name in enumerate(header_row)}
    return column_indices

def main():
    try:
        setup_logging()
        startProxy(oxyUser, oxyPW)
        prntL("starting")
        global wb
        wb = openpyxl.load_workbook(excel_file_path)
        sheet = wb.active
        column_indices = initializeIndices(sheet)
        # relevant indices (for modular use of excel sheets):
        name_index = column_indices.get("First Name")
        last_name_index = column_indices.get("Last Name")
        title_index = column_indices.get("Title")
        bibtex_index = column_indices.get("bibtex cite")
        attackDefence_index = column_indices.get("Attack/ Defense")
        pdfLink_index = column_indices.get("Article PDF link")
        isDownloaded_index= column_indices.get("isDownloaded?")

        # ~~~~~~~~~~~~~~
        # Convert the Excel sheet to a Pandas DataFrame
        data = [row for row in sheet.iter_rows(values_only=True)]
        header = [cell.value for cell in sheet[1]]
        df = pd.DataFrame(data, columns=header)

        # Specify the columns based on which you want to remove duplicates
        columns_to_check_duplicates = ["bibtex cite", "Article PDF link", "Title"]  # Adjust as needed

        # Use pandas to remove duplicates
        df_no_duplicates = df.drop_duplicates(subset=columns_to_check_duplicates, keep="first")

        # Convert the DataFrame back to an openpyxl worksheet
        sheet.delete_rows(2, sheet.max_row)
        firstimer = True
        for index, row in enumerate(df_no_duplicates.itertuples(index=False), start=2):
            if firstimer:
                firstimer = False
                continue
            sheet.append(row)


        rowNum = 1
        for row in sheet.iter_rows(min_row=2, values_only=True):
            rowNum += 1
            if not needToDownloadRow(row[isDownloaded_index],row[pdfLink_index]):
                continue
            bibtex = getBibtex(row,rowNum,name_index,last_name_index,attackDefence_index,title_index,bibtex_index,pdfLink_index)
            pdf_links, isNameDotPDF = find_pdf_links(bibtex)
            firstTimeOnly = True
            for pdf_link in pdf_links:
                if firstTimeOnly:
                    pdf_url = pdf_link['href']
                    prntL(pdf_url)
                    try:
                        if download_pdf(pdf_url, row[title_index], rowNum) == 1:
                            logging.info(f'successful downloading')
                            markAsDownloaded(True, pdf_url, rowNum, sheet,pdfLink_index,isDownloaded_index)
                        else:
                            logging.info(f'download failed')
                            markAsDownloaded(False, pdf_url, rowNum, sheet,pdfLink_index,isDownloaded_index)
                    except Exception as e:
                        logging.error(f'Exception while processing PDF on row {rowNum}: {e}')
                    finally:
                        firstTimeOnly = False
                else:
                    pdf_url = pdf_link['href']
                    prntL(f'another URL found:  {pdf_url}')
            time.sleep(1)
        prntL(f'~~FINISHED RUNNING {"with" if proxies else "WITHOUT"} proxies')
    except KeyboardInterrupt:
        prntL("Script execution manually interrupted.")
    except Exception as e:
        prntL(f'An unexpected exception occurred:\n\t {e}')

if __name__ == "__main__":
    main()
